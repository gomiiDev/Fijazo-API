"""Tests de la importación de apuestas desde Excel (plantilla, lector, servicio)."""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from fijazo_api.application.services.bet_import_service import BetImportService
from fijazo_api.core.exceptions import InvalidImportFileError
from fijazo_api.infrastructure.excel.bet_import_reader import read_bet_rows
from fijazo_api.infrastructure.excel.columns import COLUMNS
from fijazo_api.infrastructure.excel.template_generator import build_bet_template

# Naive: Excel no soporta zonas horarias en celdas de fecha (así llegan de un
# Excel real rellenado por el usuario).
EVENT_DT = datetime(2026, 8, 1, 20, 0)


def _valid_record(**overrides) -> dict:
    record = {
        "sport": "Fútbol",
        "league": "LaLiga",
        "event": "Real Madrid vs Barcelona",
        "bet_type": "SIMPLE",
        "market": "1X2",
        "selection": "Real Madrid",
        "odds": 2.0,
        "stake": 10.0,
        "bookmaker": "Bet365",
        "event_datetime": EVENT_DT,
        "status": "WON",
    }
    record.update(overrides)
    return record


def _build_xlsx(records: list[dict], *, headers: list[str] | None = None) -> bytes:
    """Crea un .xlsx con las columnas de la plantilla y las filas dadas."""

    wb = Workbook()
    ws = wb.active
    header_specs = COLUMNS
    ws.append(headers if headers is not None else [c.header for c in header_specs])
    for record in records:
        ws.append([record.get(c.field) for c in header_specs])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# --- Plantilla -------------------------------------------------------------


def test_template_headers_and_dropdowns():
    wb = load_workbook(BytesIO(build_bet_template()))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    assert headers == [c.header for c in COLUMNS]

    # Debe haber validaciones (listas) para Estado y Tipo de apuesta.
    formulas = [dv.formula1 for dv in ws.data_validations.dataValidation]
    assert any("SIMPLE" in f and "PARLAY" in f for f in formulas)
    assert any("PENDING" in f and "VOID" in f for f in formulas)


# --- Lector ----------------------------------------------------------------


def test_reader_parses_rows_and_skips_empty():
    data = _build_xlsx([_valid_record(), {}, _valid_record(reference_id="R2")])
    rows = read_bet_rows(data)

    assert len(rows) == 2  # la fila vacía se omite
    assert rows[0][0] == 2  # número de fila en Excel (fila 1 = cabecera)
    assert rows[0][1]["sport"] == "Fútbol"
    assert rows[1][0] == 4  # se saltó la fila 3 (vacía)
    assert rows[1][1]["reference_id"] == "R2"


def test_reader_missing_required_column_raises():
    headers = [c.header for c in COLUMNS if c.field != "odds"]  # falta Cuota
    data = _build_xlsx([], headers=headers)
    with pytest.raises(InvalidImportFileError):
        read_bet_rows(data)


def test_reader_rejects_non_xlsx():
    with pytest.raises(InvalidImportFileError):
        read_bet_rows(b"esto no es un excel")


# --- Servicio de importación (con fakes) -----------------------------------


class _FakeBetService:
    def __init__(self):
        self.created: list[dict] = []

    async def create_bet(self, user_id: str, data: dict):
        self.created.append(data)
        return data


class _FakeBetRepo:
    def __init__(self, existing_refs: set[str] | None = None):
        self._existing = existing_refs or set()

    async def reference_exists(self, user_id: str, reference_id: str) -> bool:
        return reference_id in self._existing


async def test_import_mixed_rows_summary():
    fake_service = _FakeBetService()
    fake_repo = _FakeBetRepo(existing_refs={"EXISTS"})
    svc = BetImportService(fake_service, fake_repo)

    rows = [
        (2, _valid_record()),  # OK
        (3, _valid_record(odds=0.5)),  # cuota inválida
        (4, _valid_record(status="INVALID")),  # estado inválido
        (5, _valid_record()),  # duplicado del de la fila 2 (evento+sel+fecha)
        (6, _valid_record(event="Otro", reference_id="R1")),  # OK con ref
        (7, _valid_record(event="Otro2", reference_id="R1")),  # ref duplicada en archivo
        (8, _valid_record(event="Otro3", reference_id="EXISTS")),  # ref ya en BD
    ]
    summary = await svc.import_rows("u1", rows)

    assert summary.total_rows == 7
    assert summary.imported == 2  # filas 2 y 6
    assert summary.rejected == 5
    assert len(fake_service.created) == 2

    fields_by_row = {(e.row, e.field) for e in summary.errors}
    assert (3, "odds") in fields_by_row
    assert (4, "status") in fields_by_row
    assert (5, "evento") in fields_by_row
    assert (7, "reference_id") in fields_by_row
    assert (8, "reference_id") in fields_by_row


async def test_import_does_not_stop_on_bad_row():
    fake_service = _FakeBetService()
    svc = BetImportService(fake_service, _FakeBetRepo())
    rows = [
        (2, _valid_record(stake=0)),  # inválida
        (3, _valid_record(event="B")),  # válida -> debe procesarse igual
    ]
    summary = await svc.import_rows("u1", rows)
    assert summary.imported == 1
    assert summary.rejected == 1


# --- Integración (API + stats) ---------------------------------------------

from httpx import AsyncClient  # noqa: E402

from tests.conftest import auth_header, register_and_login  # noqa: E402


async def test_template_endpoint(client: AsyncClient):
    token = await register_and_login(client, "tpluser", "tpl@test.com")
    resp = await client.get("/bets/template", headers=auth_header(token))
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    assert "attachment" in resp.headers["content-disposition"]
    # El contenido debe abrirse como un xlsx válido.
    assert read_bet_rows(resp.content) == []


async def test_import_endpoint_updates_stats(client: AsyncClient):
    token = await register_and_login(client, "impuser", "imp@test.com")
    h = auth_header(token)

    data = _build_xlsx(
        [
            _valid_record(status="WON", odds=2.0, stake=10),
            _valid_record(event="B", status="LOST", odds=2.0, stake=10),
            _valid_record(odds=0.5),  # inválida
        ]
    )
    files = {"file": ("apuestas.xlsx", data, "application/octet-stream")}
    resp = await client.post("/bets/import", files=files, headers=h)
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_rows"] == 3
    assert body["imported"] == 2
    assert body["rejected"] == 1
    assert body["errors"][0]["field"] == "odds"

    # Las apuestas importadas se crearon y las stats se sincronizaron.
    bets = await client.get("/bets", headers=h)
    assert bets.json()["total"] == 2
    stats = await client.get("/statistics/me", headers=h)
    s = stats.json()
    assert s["won"] == 1 and s["lost"] == 1
    assert s["win_rate"] == 50.0


async def test_import_rejects_non_xlsx(client: AsyncClient):
    token = await register_and_login(client, "baduser", "bad@test.com")
    files = {"file": ("datos.csv", b"a,b,c", "text/csv")}
    resp = await client.post("/bets/import", files=files, headers=auth_header(token))
    assert resp.status_code == 400


async def test_import_parlay_grouped_by_ticket(client: AsyncClient):
    token = await register_and_login(client, "impparlay", "impparlay@test.com")
    h = auth_header(token)

    data = _build_xlsx(
        [
            # Parlay: dos filas con el mismo Ticket "T1".
            _valid_record(bet_type="PARLAY", odds=2.0, stake=10, event="A", ticket="T1"),
            _valid_record(sport="Tenis", event="B", selection="Y", odds=3.0, ticket="T1"),
            # Una apuesta simple sin ticket.
            _valid_record(event="C"),
        ]
    )
    files = {"file": ("apuestas.xlsx", data, "application/octet-stream")}
    resp = await client.post("/bets/import", files=files, headers=h)
    assert resp.status_code == 200
    body = resp.json()
    assert body["total_rows"] == 3  # filas físicas
    assert body["imported"] == 2  # 1 parlay + 1 simple
    assert body["rejected"] == 0

    bets = (await client.get("/bets", headers=h)).json()
    assert bets["total"] == 2
    parlay = next(b for b in bets["items"] if b["bet_type"] == "PARLAY")
    assert len(parlay["legs"]) == 1
    assert parlay["combined_odds"] == 6.0  # 2.0 * 3.0
