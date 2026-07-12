"""Generación de la plantilla `.xlsx` de importación de apuestas (openpyxl)."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from fijazo_api.infrastructure.excel.columns import (
    BET_STATUS_VALUES,
    BET_TYPE_VALUES,
    COLUMNS,
)

_MAX_DATA_ROWS = 1000
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF")


def _column_letter(field: str) -> str:
    index = next(i for i, c in enumerate(COLUMNS) if c.field == field)
    return get_column_letter(index + 1)


def build_bet_template() -> bytes:
    """Construye la plantilla de importación y devuelve sus bytes `.xlsx`.

    Incluye encabezados formateados y listas desplegables para *Estado* y
    *Tipo de apuesta* (con los valores del enum).
    """

    wb = Workbook()
    ws = wb.active
    ws.title = "Apuestas"

    # Encabezados formateados + anchos.
    for col_idx, spec in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=spec.header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = spec.width

    ws.freeze_panes = "A2"

    # Listas desplegables (validación en Excel).
    for field, values in (
        ("bet_type", BET_TYPE_VALUES),
        ("status", BET_STATUS_VALUES),
    ):
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(values)}"',
            allow_blank=False,
            showDropDown=False,
        )
        letter = _column_letter(field)
        dv.add(f"{letter}2:{letter}{_MAX_DATA_ROWS + 1}")
        ws.add_data_validation(dv)

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
