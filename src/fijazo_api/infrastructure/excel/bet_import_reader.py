"""Lectura de la plantilla `.xlsx` de apuestas (openpyxl).

Aislado en infraestructura: solo lee y normaliza celdas a un dict por fila. No
aplica reglas de negocio (eso vive en la capa de aplicación).
"""

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from fijazo_api.core.exceptions import InvalidImportFileError
from fijazo_api.infrastructure.excel.columns import COLUMNS

#: Una fila leída: (número de fila en Excel, {campo: valor}). Solo incluye los
#: campos con valor; las celdas vacías se omiten para que ``BetCreate`` aplique
#: sus defaults o marque "campo requerido".
RawRow = tuple[int, dict[str, Any]]


def _normalize(value: Any) -> Any:
    if isinstance(value, str):
        value = value.strip()
        return value or None
    return value


def read_bet_rows(data: bytes) -> list[RawRow]:
    """Lee las filas de apuestas del archivo. Lanza ``InvalidImportFileError``.

    Valida que estén presentes todas las columnas obligatorias en la cabecera y
    omite las filas completamente vacías.
    """

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # archivo corrupto o que no es .xlsx
        raise InvalidImportFileError("El archivo no es un Excel (.xlsx) válido.") from exc

    ws = wb.active
    rows = ws.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        raise InvalidImportFileError("El archivo está vacío.") from None

    headers = {str(h).strip(): idx for idx, h in enumerate(header_row) if h is not None}

    missing = [c.header for c in COLUMNS if c.required and c.header not in headers]
    if missing:
        raise InvalidImportFileError("Faltan columnas obligatorias: " + ", ".join(missing))

    # Mapa campo -> índice de columna, solo para las columnas presentes.
    field_index = {c.field: headers[c.header] for c in COLUMNS if c.header in headers}

    result: list[RawRow] = []
    for offset, raw in enumerate(rows, start=2):  # la fila 1 es la cabecera
        record: dict[str, Any] = {}
        for field, idx in field_index.items():
            value = _normalize(raw[idx]) if idx < len(raw) else None
            if value is not None:
                record[field] = value
        if record:  # se omiten las filas completamente vacías
            result.append((offset, record))

    wb.close()
    return result
